"""
Text Extractor
Extract text content from various file types (PDF, images, docs)
"""
from typing import Optional
import structlog
from pathlib import Path

logger = structlog.get_logger(__name__)


class TextExtractor:
    """Extract text from various file formats"""

    def extract_text(self, file_path: str) -> Optional[str]:
        """
        Extract text from a file based on its type

        Args:
            file_path: Path to file

        Returns:
            Extracted text, or None if extraction failed
        """
        file_path = Path(file_path)

        if not file_path.exists():
            logger.error("File does not exist", file_path=str(file_path))
            return None

        # Determine file type and extract accordingly
        suffix = file_path.suffix.lower()

        try:
            if suffix == '.pdf':
                return self._extract_from_pdf(file_path)
            elif suffix in ['.jpg', '.jpeg', '.png', '.tiff', '.tif', '.bmp']:
                return self._extract_from_image(file_path)
            elif suffix == '.docx':
                return self._extract_from_docx(file_path)
            elif suffix == '.txt':
                return self._extract_from_text(file_path)
            elif suffix == '.csv':
                return self._extract_from_csv(file_path)
            elif suffix in ['.xlsx', '.xls']:
                return self._extract_from_excel(file_path)
            else:
                logger.warning("Unsupported file type for text extraction",
                             file_path=str(file_path),
                             suffix=suffix)
                return None

        except Exception as e:
            logger.error("Failed to extract text",
                        file_path=str(file_path),
                        error=str(e))
            return None

    def _extract_from_pdf(self, file_path: Path) -> Optional[str]:
        """Extract text from PDF"""
        try:
            import pdfplumber
            text_parts = []

            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)

            extracted_text = "\n\n".join(text_parts)
            logger.info("Extracted text from PDF",
                       file_path=str(file_path),
                       text_length=len(extracted_text),
                       page_count=len(text_parts))
            return extracted_text

        except ImportError:
            logger.error("pdfplumber not installed - cannot extract PDF text")
            return None
        except Exception as e:
            logger.error("Failed to extract PDF text",
                        file_path=str(file_path),
                        error=str(e))
            return None

    def _extract_from_image(self, file_path: Path) -> Optional[str]:
        """Extract text from image using OCR"""
        try:
            import pytesseract
            from PIL import Image

            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)

            logger.info("Extracted text from image using OCR",
                       file_path=str(file_path),
                       text_length=len(text))
            return text

        except ImportError:
            logger.error("pytesseract or PIL not installed - cannot perform OCR")
            return None
        except Exception as e:
            logger.error("Failed to extract text from image",
                        file_path=str(file_path),
                        error=str(e))
            return None

    def _extract_from_docx(self, file_path: Path) -> Optional[str]:
        """Extract text from Word document"""
        try:
            from docx import Document

            doc = Document(file_path)
            text_parts = []

            for paragraph in doc.paragraphs:
                if paragraph.text:
                    text_parts.append(paragraph.text)

            extracted_text = "\n".join(text_parts)
            logger.info("Extracted text from DOCX",
                       file_path=str(file_path),
                       text_length=len(extracted_text),
                       paragraph_count=len(text_parts))
            return extracted_text

        except ImportError:
            logger.error("python-docx not installed - cannot extract DOCX text")
            return None
        except Exception as e:
            logger.error("Failed to extract DOCX text",
                        file_path=str(file_path),
                        error=str(e))
            return None

    def _extract_from_text(self, file_path: Path) -> Optional[str]:
        """Extract text from plain text file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()

            logger.info("Extracted text from text file",
                       file_path=str(file_path),
                       text_length=len(text))
            return text

        except Exception as e:
            logger.error("Failed to read text file",
                        file_path=str(file_path),
                        error=str(e))
            return None

    def _extract_from_csv(self, file_path: Path) -> Optional[str]:
        """Extract text from CSV file"""
        try:
            import csv

            rows = []
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(" | ".join(row))

            extracted_text = "\n".join(rows)
            logger.info("Extracted text from CSV",
                       file_path=str(file_path),
                       text_length=len(extracted_text),
                       row_count=len(rows))
            return extracted_text

        except Exception as e:
            logger.error("Failed to extract CSV text",
                        file_path=str(file_path),
                        error=str(e))
            return None

    def _extract_from_excel(self, file_path: Path) -> Optional[str]:
        """Extract text from Excel file"""
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"=== Sheet: {sheet_name} ===")

                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)

            extracted_text = "\n".join(text_parts)
            logger.info("Extracted text from Excel",
                       file_path=str(file_path),
                       text_length=len(extracted_text),
                       sheet_count=len(workbook.sheetnames))
            return extracted_text

        except ImportError:
            logger.error("openpyxl not installed - cannot extract Excel text")
            return None
        except Exception as e:
            logger.error("Failed to extract Excel text",
                        file_path=str(file_path),
                        error=str(e))
            return None
